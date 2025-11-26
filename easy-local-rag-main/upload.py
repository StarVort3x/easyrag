import os
import tkinter as tk
from tkinter import filedialog, messagebox
import PyPDF2
import re
import json
from openpyxl import load_workbook

# Function to convert PDF to text and append to vault.txt
def convert_pdf_to_text():
    file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    if file_path:
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
            text = ''
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                if page.extract_text():
                    text += page.extract_text() + " "
            
            # Normalize whitespace and clean up text
            text = re.sub(r'\s+', ' ', text).strip()
            
            # Split text into chunks by sentences, respecting a maximum chunk size
            sentences = re.split(r'(?<=[.!?]) +', text)  # split on spaces following sentence-ending punctuation
            chunks = []
            current_chunk = ""
            for sentence in sentences:
                # Check if the current sentence plus the current chunk exceeds the limit
                if len(current_chunk) + len(sentence) + 1 < 1000:  # +1 for the space
                    current_chunk += (sentence + " ").strip()
                else:
                    # When the chunk exceeds 1000 characters, store it and start a new one
                    chunks.append(current_chunk)
                    current_chunk = sentence + " "
            if current_chunk:  # Don't forget the last chunk!
                chunks.append(current_chunk)
            with open("vault.txt", "a", encoding="utf-8") as vault_file:
                for chunk in chunks:
                    # Write each chunk to its own line
                    vault_file.write(chunk.strip() + "\n")  # Two newlines to separate chunks
            print(f"PDF content appended to vault.txt with each chunk on a separate line.")

# Function to upload a text file and append to vault.txt
def upload_txtfile():
    file_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
    if file_path:
        with open(file_path, 'r', encoding="utf-8") as txt_file:
            text = txt_file.read()
            
            # Normalize whitespace and clean up text
            text = re.sub(r'\s+', ' ', text).strip()
            
            # Split text into chunks by sentences, respecting a maximum chunk size
            sentences = re.split(r'(?<=[.!?]) +', text)  # split on spaces following sentence-ending punctuation
            chunks = []
            current_chunk = ""
            for sentence in sentences:
                # Check if the current sentence plus the current chunk exceeds the limit
                if len(current_chunk) + len(sentence) + 1 < 1000:  # +1 for the space
                    current_chunk += (sentence + " ").strip()
                else:
                    # When the chunk exceeds 1000 characters, store it and start a new one
                    chunks.append(current_chunk)
                    current_chunk = sentence + " "
            if current_chunk:  # Don't forget the last chunk!
                chunks.append(current_chunk)
            with open("vault.txt", "a", encoding="utf-8") as vault_file:
                for chunk in chunks:
                    # Write each chunk to its own line
                    vault_file.write(chunk.strip() + "\n")  # Two newlines to separate chunks
            print(f"Text file content appended to vault.txt with each chunk on a separate line.")

# Function to upload a JSON file and append to vault.txt
def upload_jsonfile():
    file_path = filedialog.askopenfilename(filetypes=[("JSON Files", "*.json")])
    if file_path:
        with open(file_path, 'r', encoding="utf-8") as json_file:
            data = json.load(json_file)
            
            # Flatten the JSON data into a single string
            text = json.dumps(data, ensure_ascii=False)
            
            # Normalize whitespace and clean up text
            text = re.sub(r'\s+', ' ', text).strip()
            
            # Split text into chunks by sentences, respecting a maximum chunk size
            sentences = re.split(r'(?<=[.!?]) +', text)  # split on spaces following sentence-ending punctuation
            chunks = []
            current_chunk = ""
            for sentence in sentences:
                # Check if the current sentence plus the current chunk exceeds the limit
                if len(current_chunk) + len(sentence) + 1 < 1000:  # +1 for the space
                    current_chunk += (sentence + " ").strip()
                else:
                    # When the chunk exceeds 1000 characters, store it and start a new one
                    chunks.append(current_chunk)
                    current_chunk = sentence + " "
            if current_chunk:  # Don't forget the last chunk!
                chunks.append(current_chunk)
            with open("vault.txt", "a", encoding="utf-8") as vault_file:
                for chunk in chunks:
                    # Write each chunk to its own line
                    vault_file.write(chunk.strip() + "\n")  # Two newlines to separate chunks
            print(f"JSON file content appended to vault.txt with each chunk on a separate line.")

def upload_excelfile():
    file_path = filedialog.askopenfilename(
        filetypes=[
            ("Excel Files", "*.xlsx *.xls"),
            ("All Files", "*.*")
        ]
    )
    
    if not file_path:
        return
        
    try:
        # 根据文件扩展名选择不同的处理方式
        if file_path.lower().endswith('.xls'):
            # 处理 .xls 文件
            import xlrd
            wb = xlrd.open_workbook(file_path)
            text_chunks = []
            
            # 处理每个工作表
            for sheet_name in wb.sheet_names():
                sheet = wb.sheet_by_name(sheet_name)
                sheet_text = []
                
                # 添加工作表标题
                sheet_text.append(f"## {sheet_name}\n")
                
                # 获取表格数据
                table_data = []
                max_cols = 0
                
                # 读取所有行
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        # 处理不同类型的单元格值
                        if isinstance(cell_value, (int, float)):
                            cell_value = str(cell_value)
                        elif cell_value is None:
                            cell_value = ""
                        else:
                            cell_value = str(cell_value).strip()
                        row_data.append(cell_value)
                    
                    # 更新最大列数
                    max_cols = max(max_cols, len(row_data))
                    table_data.append(row_data)
                
                # 如果没有数据，跳过该工作表
                if not table_data:
                    continue
                
                # 创建Markdown表格
                md_table = []
                
                # 添加表头
                if table_data:
                    header = table_data[0]
                    md_table.append("| " + " | ".join(header) + " |")
                    # 添加分隔行
                    md_table.append("| " + " | ".join(["---"] * len(header)) + " |")
                    
                    # 添加数据行
                    for row in table_data[1:]:
                        # 确保每行的列数与表头一致
                        row = row + [""] * (len(header) - len(row))
                        md_table.append("| " + " | ".join(row) + " |")
                
                # 将表格添加到结果中
                if md_table:
                    sheet_text.append("\n".join(md_table))
                    text_chunks.append("\n".join(sheet_text))
            
        else:
            # 处理 .xlsx 文件
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            text_chunks = []
            
            # 处理每个工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_text = []
                
                # 添加工作表标题
                sheet_text.append(f"## {sheet_name}\n")
                
                # 获取表格数据
                table_data = []
                max_cols = 0
                
                # 读取所有行
                for row in ws.iter_rows(values_only=True):
                    row_data = []
                    for cell in row:
                        # 处理不同类型的单元格值
                        if cell is None:
                            cell_value = ""
                        elif isinstance(cell, (int, float)):
                            cell_value = str(cell)
                        else:
                            cell_value = str(cell).strip()
                        row_data.append(cell_value)
                    
                    # 更新最大列数
                    max_cols = max(max_cols, len(row_data))
                    table_data.append(row_data)
                
                # 如果没有数据，跳过该工作表
                if not table_data:
                    continue
                
                # 创建Markdown表格
                md_table = []
                
                # 添加表头
                if table_data:
                    header = table_data[0]
                    # 确保表头不为空
                    if not any(header):
                        header = [f"列{idx+1}" for idx in range(len(header))]
                    md_table.append("| " + " | ".join(header) + " |")
                    # 添加分隔行
                    md_table.append("| " + " | ".join(["---"] * len(header)) + " |")
                    
                    # 添加数据行
                    for row in table_data[1:]:
                        # 确保每行的列数与表头一致
                        row = list(row) if row is not None else []
                        row = row + [""] * (len(header) - len(row))
                        # 处理每列数据
                        formatted_row = []
                        for cell in row:
                            if cell is None:
                                cell = ""
                            elif isinstance(cell, (int, float)):
                                cell = str(cell)
                            else:
                                cell = str(cell).strip()
                            # 转义Markdown特殊字符
                            cell = cell.replace("|", "\\|").replace("\n", "<br>")
                            formatted_row.append(cell)
                        md_table.append("| " + " | ".join(formatted_row) + " |")
                
                # 将表格添加到结果中
                if md_table:
                    sheet_text.append("\n".join(md_table))
                    text_chunks.append("\n".join(sheet_text))
        
        if not text_chunks:  # 如果没有提取到任何内容
            messagebox.showwarning("警告", "Excel文件为空或没有可读数据。")
            return
        
        # 合并所有文本
        full_text = "\n\n".join(text_chunks)
        
        # 写入vault文件
        with open("vault.txt", "a", encoding="utf-8") as vault_file:
            vault_file.write("\n\n" + full_text + "\n\n")
        
        messagebox.showinfo("成功", "Excel文件内容已处理并添加到vault.txt")
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        messagebox.showerror("错误", f"处理Excel文件时出错：\n{str(e)}\n\n详细信息：\n{error_details}")
    finally:
        if 'wb' in locals():
            # xlrd 的 Workbook 对象没有 close 方法
            if not isinstance(wb, type(None)) and hasattr(wb, 'release_resources'):
                wb.release_resources()
            elif hasattr(wb, 'close'):
                wb.close()

# Create the main window
root = tk.Tk()
root.title("Upload .pdf, .txt, or .json")

# Create buttons for each function
pdf_button = tk.Button(root, text="Upload PDF", command=convert_pdf_to_text)
txt_button = tk.Button(root, text="Upload Text File", command=upload_txtfile)
json_button = tk.Button(root, text="Upload JSON File", command=upload_jsonfile)
excel_button = tk.Button(root, text="Upload Excel File", command=upload_excelfile)
    
# Pack buttons
pdf_button.pack(pady=5)
txt_button.pack(pady=5)
json_button.pack(pady=5)
excel_button.pack(pady=5)

# Run the main event loop
root.mainloop()
